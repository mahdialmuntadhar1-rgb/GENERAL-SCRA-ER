"""
Import CSV / Excel / JSON files into a list[dict] for staging.
Uses only stdlib csv + openpyxl (no pandas/numpy dependency).
"""

import csv
import json
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook


SUPPORTED_EXTENSIONS = {".csv", ".xlsx", ".xls", ".json"}


def load_file(path: str) -> list[dict]:
    """
    Load a data file and return a list of dicts (one per row/record).
    Supports .csv, .xlsx/.xls, and .json.
    Raises ValueError for unsupported formats.
    """
    p = Path(path)
    ext = p.suffix.lower()

    if ext == ".csv":
        return _load_csv(p)
    elif ext in (".xlsx", ".xls"):
        return _load_excel(p)
    elif ext == ".json":
        return _load_json(p)
    else:
        raise ValueError(
            f"Unsupported file type: {ext}. "
            f"Supported: {', '.join(sorted(SUPPORTED_EXTENSIONS))}"
        )


def _normalise_header(name: str) -> str:
    return name.strip().lower().replace(" ", "_")


def _load_csv(path: Path) -> list[dict]:
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        if reader.fieldnames is None:
            return []
        headers = [_normalise_header(h) for h in reader.fieldnames]
        rows: list[dict] = []
        for raw_row in reader:
            row = {
                _normalise_header(k): (str(v).strip() if v else "")
                for k, v in raw_row.items()
            }
            rows.append(row)
    return rows


def _load_excel(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return []
    row_iter = ws.iter_rows(values_only=True)
    try:
        raw_headers = next(row_iter)
    except StopIteration:
        return []
    headers = [_normalise_header(str(h) if h else f"col_{i}") for i, h in enumerate(raw_headers)]
    rows: list[dict] = []
    for raw_row in row_iter:
        row = {
            headers[i]: (str(cell).strip() if cell is not None else "")
            for i, cell in enumerate(raw_row)
            if i < len(headers)
        }
        rows.append(row)
    wb.close()
    return rows


def _load_json(path: Path) -> list[dict]:
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    if isinstance(data, list):
        return data
    elif isinstance(data, dict):
        # Assume single record
        return [data]
    else:
        raise ValueError("JSON must be an array of objects or a single object")


def detect_column_mapping(columns: list[str]) -> dict[str, Optional[str]]:
    """
    Heuristically map incoming column names to our schema fields.
    Returns {schema_field: source_column_or_None}.
    """
    ALIASES: dict[str, list[str]] = {
        "name":           ["name", "university_name", "uni_name", "institution", "اسم"],
        "name_ar":        ["name_ar", "arabic_name", "الاسم_العربي", "اسم_عربي"],
        "name_en":        ["name_en", "english_name"],
        "type":           ["type", "uni_type", "university_type", "نوع"],
        "address":        ["address", "location", "عنوان"],
        "city":           ["city", "city_name", "مدينة"],
        "governorate":    ["governorate", "gov", "province", "محافظة"],
        "latitude":       ["latitude", "lat"],
        "longitude":      ["longitude", "lon", "lng"],
        "founded_year":   ["founded_year", "founded", "year", "سنة_التأسيس"],
        "website":        ["website", "url", "web", "موقع"],
        "phone":          ["phone", "telephone", "tel", "هاتف"],
        "email":          ["email", "mail", "بريد"],
        "facebook":       ["facebook", "fb"],
        "instagram":      ["instagram", "ig"],
        "description":    ["description", "desc", "about", "وصف"],
    }

    lower_cols = {c.lower(): c for c in columns}
    mapping: dict[str, Optional[str]] = {}

    for field, aliases in ALIASES.items():
        matched = None
        for alias in aliases:
            if alias in lower_cols:
                matched = lower_cols[alias]
                break
        mapping[field] = matched

    return mapping
